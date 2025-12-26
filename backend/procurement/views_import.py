"""
Импорт прайс‑листов в procurement.

Реализует загрузку Excel‑файла прайса от поставщика и преобразование строк в PriceRecord.
Поддерживает режим preview, чтобы пользователь увидел результат сопоставления до сохранения.
"""

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction
from catalog.models import Item
from suppliers.models import Supplier
from procurement.models import PriceRecord
from datetime import date, datetime as dtmod
import openpyxl


from .importers._resolver import resolve_item_id_by_supplier_context


EXPECTED_HEADERS = {
    "item_sku":"item_sku", "supplier":"supplier", "price":"price", "currency":"currency",
    "lead_days":"lead_days", "pack_qty":"pack_qty", "moq_qty":"moq_qty", "mo_amount":"mo_amount",
    "lot_step":"lot_step", "vat_included":"vat_included", "vat_rate":"vat_rate",
    "delivery_fixed":"delivery_fixed", "delivery_per_unit":"delivery_per_unit", "dt":"dt",
    "номенклатура":"item_sku", "поставщик":"supplier", "цена":"price", "валюта":"currency",
    "срок поставки (дн.)":"lead_days", "кратность упаковки":"pack_qty",
    "мин. кол-во (moq)":"moq_qty", "мин. сумма заказа":"mo_amount", "кратность заказа поставщика":"lot_step",
    "цена с ндс":"vat_included", "ставка ндс, %":"vat_rate", "доставка фикс.":"delivery_fixed",
    "доставка за ед.":"delivery_per_unit", "дата":"dt",
}


def _lower_map(ws):
    """
    Функция _lower_map относится к бизнес‑логике SNAB. См. код и параметры вызова.
    """

    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v is None: continue
        key = str(v).strip().lower()
        if key in EXPECTED_HEADERS:
            headers[EXPECTED_HEADERS[key]] = col
    return headers


def _to_bool(val, default=None):
    """
    Функция _to_bool относится к бизнес‑логике SNAB. См. код и параметры вызова.
    """

    if val is None: return default
    if isinstance(val, bool): return val
    s = str(val).strip().lower()
    if s in ("1","true","yes","да","y","истина"): return True
    if s in ("0","false","no","нет","n","ложь"): return False
    return default


def _to_decimal(val, default=None):
    """
    Функция _to_decimal относится к бизнес‑логике SNAB. См. код и параметры вызова.
    """

    if val in (None, ""): return default
    try: return float(val)
    except: return default


def _to_int(val, default=None):
    """
    Функция _to_int относится к бизнес‑логике SNAB. См. код и параметры вызова.
    """

    if val in (None, ""): return default
    try: return int(val)
    except: return default


def _to_date(val, default=None):
    """
    Функция _to_date относится к бизнес‑логике SNAB. См. код и параметры вызова.
    """

    if val in (None, ""): return default
    if isinstance(val, date): return val
    if isinstance(val, dtmod): return val.date()
    try: return date.fromisoformat(str(val)[:10])
    except: return default


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_price_excel(request):
    """
    Импорт прайс‑листа поставщика из Excel (openpyxl).
    Поддерживает режим preview=1: вернуть распознанные строки без записи в БД.
    Ожидает файл в multipart/form-data под ключом 'file'.
    Сопоставление номенклатуры выполняется через resolver (по артикулу/контексту поставщика).
    """

    preview = request.query_params.get("preview") in ("1","true","True")
    file = request.FILES.get("file")
    if not file:
        return Response({"detail":"Нет файла"}, status=400)
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return Response({"detail": f"Excel error: {e}"}, status=400)


    headers = _lower_map(ws)
    required = ["item_sku", "supplier", "price"]
    missing = [h for h in required if h not in headers]
    if missing:
        return Response({"detail": f"Missing columns: {missing}"}, status=400)


    preview_rows = []
    valid_rows = []
    errors = 0


    for r in range(2, ws.max_row + 1):
        sku_raw = ws.cell(r, headers["item_sku"]).value
        if sku_raw in (None, ""): 
            continue
        supplier_val = ws.cell(r, headers["supplier"]).value if "supplier" in headers else None
        price = _to_decimal(ws.cell(r, headers["price"]).value)
        currency = (ws.cell(r, headers["currency"]).value if "currency" in headers else "RUB") or "RUB"
        supplier_name = str(supplier_val).strip() if supplier_val else ""


        row_err = {}
        if price is None:
            row_err["price"] = "invalid"


        sku = str(sku_raw).strip()
        item_id = resolve_item_id_by_supplier_context(supplier_name, sku)


        candidates = []
        if not item_id:
            row_err["item_sku"] = f"unmapped: {sku}"
            candidates = list(Item.objects.filter(sku__icontains=sku).values_list('id','sku')[:5])


        lead_days = _to_int(ws.cell(r, headers.get("lead_days", 0)).value, 0) if "lead_days" in headers else 0
        pack_qty = _to_decimal(ws.cell(r, headers.get("pack_qty", 0)).value, 1) if "pack_qty" in headers else 1
        moq_qty = _to_decimal(ws.cell(r, headers.get("moq_qty", 0)).value, 0) if "moq_qty" in headers else 0
        mo_amount = _to_decimal(ws.cell(r, headers.get("mo_amount", 0)).value, 0) if "mo_amount" in headers else 0
        lot_step = _to_decimal(ws.cell(r, headers.get("lot_step", 0)).value, 1) if "lot_step" in headers else 1
        vat_included = _to_bool(ws.cell(r, headers.get("vat_included", 0)).value, True) if "vat_included" in headers else True
        vat_rate = _to_decimal(ws.cell(r, headers.get("vat_rate", 0)).value, 20) if "vat_rate" in headers else 20
        delivery_fixed = _to_decimal(ws.cell(r, headers.get("delivery_fixed", 0)).value, 0) if "delivery_fixed" in headers else 0
        delivery_per_unit = _to_decimal(ws.cell(r, headers.get("delivery_per_unit", 0)).value, 0) if "delivery_per_unit" in headers else 0
        dt_val = _to_date(ws.cell(r, headers.get("dt", 0)).value, date.today()) if "dt" in headers else date.today()


        row_out = {
            "row": r,
            "valid": len(row_err) == 0,
            "errors": row_err or None,
            "supplier": supplier_name,
            "item_sku": sku,
            "item": item_id,
            "candidates": [{"id": c[0], "sku": c[1]} for c in candidates],
            "price": price,
            "currency": currency,
            "lead_days": lead_days,
            "pack_qty": pack_qty,
            "moq_qty": moq_qty,
            "mo_amount": mo_amount,
            "lot_step": lot_step,
            "vat_included": vat_included,
            "vat_rate": vat_rate,
            "delivery_fixed": delivery_fixed,
            "delivery_per_unit": delivery_per_unit,
            "dt": dt_val.isoformat(),
        }
        preview_rows.append(row_out)
        if row_out["valid"]:
            valid_rows.append(row_out)
        else:
            errors += 1


    if preview:
        return Response({"ok": True, "preview": True, "rows": preview_rows, "errors": errors}, status=200)


    created = 0
    with transaction.atomic():
        for r in valid_rows:
            PriceRecord.objects.create(
                item_id=r["item"],
                supplier_id=Supplier.objects.only('id').get(name=r["supplier"]).id if r["supplier"] else None,
                price=r["price"],
                currency=r["currency"],
                dt=r["dt"],
                pack_qty=r["pack_qty"],
                lead_days=r["lead_days"],
                moq_qty=r["moq_qty"],
                lot_step=r["lot_step"],
            )
            created += 1
    return Response({"ok": True, "preview": False, "created": created, "errors": errors}, status=200)