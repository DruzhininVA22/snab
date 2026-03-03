"""
views.py — вьюхи для модуля закупок (procurement) и управления прайс-листами.

Реализовано:
- PriceRecordViewSet — read-only история цен,
- PurchaseRequestViewSet — CRUD заявок с nested строками и /refs/.
- SupplierPriceListViewSet — CRUD для прайс-листов + upload
- SupplierPriceListLineViewSet — CRUD для позиций
* ItemSupplierMappingViewSet — CRUD для сопоставлений + поиск альтернатив

Плюс вспомогательные функции:
- generate_quotes_from_request() — автогенерация КП из заявки
- import_price_excel() — загрузка прайс-листа из файла (из views_import.py)
"""

from django.db import transaction
from django.db.models import Q, F, Prefetch
from django.utils import timezone

from rest_framework import viewsets, permissions, status
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action, api_view
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.views import APIView

from .models import (
    PriceRecord, 
    PurchaseRequest, 
    PurchaseRequestLine,
    PurchaseOrder, 
    PurchaseOrderLine, 
    Quote, 
    QuoteLine,
    SupplierPriceList, 
    SupplierPriceListLine, 
    ItemSupplierMapping,
)

from .serializers import (
    PriceRecordSerializer,
    PurchaseRequestSerializer,
    PurchaseRequestWriteSerializer,
    PurchaseOrderSerializer,
    PurchaseOrderLineSerializer,
    PurchaseRequestLineSerializer,
    QuoteSerializer,
    ShipmentSerializer,
    SupplierPriceListSerializer,
    SupplierPriceListDetailSerializer,
    SupplierPriceListLineSerializer,
    SupplierPriceListLineDetailSerializer,
    ItemSupplierMappingSerializer,
    ItemSupplierOptionsSerializer,
    GenerateQuotesFromRequestSerializer,
    GenerateQuotesResponseSerializer,)

from projects.models import Project, ProjectStage
from suppliers.models import Supplier
from datetime import date, timedelta, datetime
from collections import defaultdict
from decimal import Decimal
import openpyxl

# --- PR status recalculation rules ---
# "Обеспечение" заявки считаем ТОЛЬКО по заказам, которые реально отправлены поставщику (sent) и дальше по цепочке.
# Закрываем заявку ТОЛЬКО когда всё покрыто заказами со статусом confirmed (и дальше).
PO_ORDERED_STATUSES = {"sent", "confirmed", "paid", "in_transit", "delivered", "closed"}
PO_CONFIRMED_STATUSES = {"confirmed", "paid", "in_transit", "delivered", "closed"}


def _pr_required_qty_by_item(pr: PurchaseRequest) -> dict:
    req: dict[int, Decimal] = {}
    for ln in pr.lines.all():
        if ln.item_id is None:
            continue
        req[ln.item_id] = req.get(ln.item_id, Decimal("0")) + (ln.qty or Decimal("0"))
    return req


def _pr_ordered_qty_by_item(pr: PurchaseRequest, statuses: set[str], exclude_po_id: int | None = None) -> dict:
    q = PurchaseOrderLine.objects.filter(order__purchase_request=pr, order__status__in=statuses)
    if exclude_po_id:
        q = q.exclude(order_id=exclude_po_id)
    out: dict[int, Decimal] = {}
    for ln in q:
        if ln.item_id is None:
            continue
        out[ln.item_id] = out.get(ln.item_id, Decimal("0")) + (ln.qty or Decimal("0"))
    return out


def _is_fully_covered(required: dict, covered: dict) -> bool:
    for item_id, need_qty in required.items():
        got = covered.get(item_id, Decimal("0"))
        if got != need_qty:
            return False
    return True


def _validate_po_not_overorder(po: PurchaseOrder, new_status: str) -> None:
    """Запрещаем перевести заказ в sent/confirmed если он приводит к перезаказу по заявке.

    Правило:
    - сравниваем по item_id (агрегированно), т.к. в PO lines нет ссылки на конкретную строку заявки.
    - проверяем только при переводе в статус из PO_ORDERED_STATUSES/PO_CONFIRMED_STATUSES.
    """
    pr = po.purchase_request
    if pr is None:
        return

    required = _pr_required_qty_by_item(pr)
    if not required:
        return

    already = _pr_ordered_qty_by_item(pr, PO_ORDERED_STATUSES, exclude_po_id=po.id)

    errors: list[str] = []
    for ln in po.lines.all():
        item_id = ln.item_id
        if item_id is None:
            continue

        if item_id not in required:
            errors.append(f"Товар item_id={item_id} отсутствует в заявке PR#{pr.id}.")
            continue

        need = required[item_id]
        have = already.get(item_id, Decimal("0"))
        after = have + (ln.qty or Decimal("0"))

        if after > need:
            errors.append(
                f"Перезаказ по item_id={item_id}: нужно {need}, уже отправлено {have}, в этом заказе {ln.qty} → будет {after}."
            )

    if errors:
        from rest_framework.exceptions import ValidationError
        raise ValidationError({"status": [f"Нельзя перевести заказ в '{new_status}'."] + errors})


def _recalc_purchase_request_status(pr: PurchaseRequest) -> None:
    # Не трогаем отменённые вручную заявки
    if (pr.status or "").lower() in {"cancelled", "canceled"}:
        return

    required = _pr_required_qty_by_item(pr)
    if not required:
        return

    ordered = _pr_ordered_qty_by_item(pr, PO_ORDERED_STATUSES)
    confirmed = _pr_ordered_qty_by_item(pr, PO_CONFIRMED_STATUSES)

    if _is_fully_covered(required, confirmed):
        new_status = "closed"
    elif _is_fully_covered(required, ordered):
        new_status = "open"
    else:
        new_status = "draft"

    if pr.status != new_status:
        pr.status = new_status
        pr.save(update_fields=["status"])


class SafeReadOnlyMixin:
    """
    Миксин, копирующий старое поведение:
    - SAFE_METHODS (GET/HEAD/OPTIONS) доступны всем,
    - изменяющие методы требуют аутентификацию.
    """

    def get_permissions(self):
        method = getattr(self.request, "method", "GET")
        if method in permissions.SAFE_METHODS:
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated()]

class PriceRecordViewSet(viewsets.ReadOnlyModelViewSet):
    """
    /api/procurement/pricerecords/

    Только чтение истории цен.
    """

    queryset = PriceRecord.objects.all().select_related("supplier", "item")
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PriceRecordSerializer


class PurchaseRequestViewSet(viewsets.ModelViewSet):
    """
    /api/procurement/purchase-requests/

    ViewSet для заявок на закупку (PurchaseRequest) c nested строками.
    """

    queryset = PurchaseRequest.objects.all().prefetch_related("lines")
    permission_classes = [permissions.IsAuthenticated]

    def get_serializer_class(self):
        if self.action in ("create", "update", "partial_update"):
            return PurchaseRequestWriteSerializer
        return PurchaseRequestSerializer

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        serializer = PurchaseRequestWriteSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        pr = serializer.save()
        out = PurchaseRequestSerializer(pr, context={"request": request})
        return Response(out.data, status=status.HTTP_201_CREATED)

    @transaction.atomic
    def update(self, request, *args, **kwargs):
        partial = kwargs.get("partial", False)
        instance = self.get_object()
        serializer = PurchaseRequestWriteSerializer(
            instance, data=request.data, partial=partial
        )
        serializer.is_valid(raise_exception=True)
        pr = serializer.save()
        out = PurchaseRequestSerializer(pr, context={"request": request})
        return Response(out.data)

    @transaction.atomic
    def partial_update(self, request, *args, **kwargs):
        kwargs["partial"] = True
        return self.update(request, *args, **kwargs)

    @action(methods=["get"], detail=False, url_path="refs")
    def refs(self, request, *args, **kwargs):
        """
        GET /api/procurement/purchase-requests/refs/

        Возвращает справочные данные:
        - список проектов,
        - список этапов проектов.
        """
        projects = list(
            Project.objects.all()
            .order_by("name")
            .values("id", "code", "name", "status")
        )
        stages = list(
            ProjectStage.objects.all()
            .order_by("project_id", "order")
            .values(
                "id",
                "project_id",
                "name",
                "status",
                "planned_start",
                "planned_end",
            )
        )
        return Response({"projects": projects, "stages": stages})

class PurchaseOrderViewSet(viewsets.ModelViewSet):
    """
    /api/procurement/purchase-orders/

    CRUD по заказам поставщикам.
    """

    queryset = (
        PurchaseOrder.objects
        .select_related("supplier")
        .prefetch_related("lines")
        .order_by("-id")
    )
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PurchaseOrderSerializer

    def partial_update(self, request, *args, **kwargs):
        instance: PurchaseOrder = self.get_object()
        old_status = (instance.status or "").lower()
        new_status = (request.data.get("status") or "").lower().strip()

        # Дедлайн заказа редактируем только в черновике
        if "deadline" in request.data and old_status != "draft":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Дедлайн можно менять только в статусе 'Черновик'.")

        # Если переводим в "sent/confirmed/..." — проверяем, что не перезаказываем относительно заявки.
        if new_status and new_status != old_status and new_status in PO_ORDERED_STATUSES.union(PO_CONFIRMED_STATUSES):
            _validate_po_not_overorder(instance, new_status)

        resp = super().partial_update(request, *args, **kwargs)

        # После успешного обновления статуса — пересчитываем статус заявки.
        try:
            instance.refresh_from_db()
            if instance.purchase_request_id:
                _recalc_purchase_request_status(instance.purchase_request)
        except Exception:
            # Не ломаем ответ из-за пересчёта статуса (логирование можно добавить позже)
            pass

        return resp


    @action(detail=True, methods=["post"])
    def send(self, request, pk=None):
        """Отправить заказ поставщику (минимальный compat endpoint).

        URL: POST /api/procurement/purchase-orders/<id>/send/

        Правило:
        - отправлять можно только из draft
        - перед отправкой обязательны delivery_address и planned_delivery_date
        """
        po: PurchaseOrder = self.get_object()

        # Идемпотентность: если уже не draft — просто вернём текущий заказ
        if (po.status or "").lower() != "draft":
            ser = self.get_serializer(po)
            return Response(ser.data)

        if not (po.delivery_address or "").strip():
            return Response({"detail": "Укажите адрес доставки перед отправкой заказа."}, status=status.HTTP_400_BAD_REQUEST)

        if not getattr(po, "planned_delivery_date", None):
            return Response({"detail": "Укажите планируемую дату поставки перед отправкой заказа."}, status=status.HTTP_400_BAD_REQUEST)

        po.status = "sent"
        po.sent_at = timezone.now()
        po.save(update_fields=["status", "sent_at"])

        # Пересчёт статуса заявки (если есть)
        try:
            if po.purchase_request_id:
                _recalc_purchase_request_status(po.purchase_request)
        except Exception:
            pass

        ser = self.get_serializer(po)
        return Response(ser.data)


    def destroy(self, request, *args, **kwargs):
        instance: PurchaseOrder = self.get_object()
        pr = instance.purchase_request
        resp = super().destroy(request, *args, **kwargs)
        try:
            if pr is not None:
                _recalc_purchase_request_status(pr)
        except Exception:
            pass
        return resp



class PurchaseOrderLineViewSet(viewsets.ModelViewSet):
    """CRUD для строк заказов.

    Ограничение:
    - редактирование qty/price допускается ТОЛЬКО пока заказ в статусе 'draft'.
    """

    queryset = PurchaseOrderLine.objects.select_related("order", "item").order_by("-id")
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = PurchaseOrderLineSerializer

    def partial_update(self, request, *args, **kwargs):
        instance: PurchaseOrderLine = self.get_object()
        if (instance.order.status or "").lower() != "draft":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Редактировать строки можно только в статусе 'Черновик'.")
        return super().partial_update(request, *args, **kwargs)

    def update(self, request, *args, **kwargs):
        instance: PurchaseOrderLine = self.get_object()
        if (instance.order.status or "").lower() != "draft":
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied("Редактировать строки можно только в статусе 'Черновик'.")
        return super().update(request, *args, **kwargs)


@api_view(["GET"])
def metrics_overview(request):
    """
    GET /api/procurement/metrics-overview/

    Краткие метрики для дашборда.
    """
    # заявки
    total_pr = PurchaseRequest.objects.count()
    open_pr = PurchaseRequest.objects.filter(status="open").count()

    # строки заявок в ожидании
    waiting_lines = PurchaseRequestLine.objects.filter(status="pending").count()

    # простая раскладка по статусам сроков (ETA/SLAs) — можно доработать позже
    today = date.today()
    warn_date = today + timedelta(days=3)

    sla = defaultdict(int)
    qs = PurchaseRequestLine.objects.all().only("need_date", "status")
    for ln in qs:
        if not ln.need_date:
            continue
        if ln.need_date <= today:
            sla["hot"] += 1
        elif ln.need_date <= warn_date:
            sla["warn"] += 1
        else:
            sla["ok"] += 1

    data = {
        "purchase_requests_total": total_pr,
        "purchase_requests_open": open_pr,
        "lines_waiting": waiting_lines,
        "sla_hot": sla["hot"],
        "sla_warn": sla["warn"],
        "sla_ok": sla["ok"],
    }

    return Response(data)

class QuoteViewSet(viewsets.ModelViewSet):
    """
    API для коммерческих предложений.

    Маршруты:
    - GET /api/procurement/quotes/        — список КП
    - GET /api/procurement/quotes/{id}/   — детали КП
    - PATCH /api/procurement/quotes/{id}/ — смена статуса/notes
    - POST /api/procurement/quotes/{id}/create_po/ — создать заказ из КП
    """

    queryset = Quote.objects.all().select_related("supplier").prefetch_related("lines__item")
    permission_classes = [permissions.IsAuthenticated]
    serializer_class = QuoteSerializer









    def get_queryset(self):
        qs = super().get_queryset()
        pr_id = (
            self.request.query_params.get("purchase_request_id")
            or self.request.query_params.get("purchase_request")
            or self.request.query_params.get("pr_id")
        )
        if pr_id:
            try:
                qs = qs.filter(purchase_request_id=int(pr_id))
            except Exception:
                pass
        return qs

    def partial_update(self, request, *args, **kwargs):
        """
        Разрешаем частичное обновление статуса и примечаний.
        """
        instance = self.get_object()
        status_val = request.data.get("status")
        notes_val = request.data.get("notes")

        # В модели Quote пока нет полей status/notes — сохраняем в cache-based метаданных.
        from procurement.serializers import _quote_meta_set
        _quote_meta_set(instance.id, status=status_val, notes=notes_val)

        serializer = self.get_serializer(instance)
        return Response(serializer.data)

    @action(detail=True, methods=["post"])
    @transaction.atomic
    def create_po(self, request, pk=None):
        """
        Создаёт PurchaseOrder на основе выбранного КП.

        Логика:
        - supplier — как в Quote;
        - status — draft (до отправки поставщику);
        - purchase_request/quote — проставляем, чтобы дальше можно было считать обеспечение заявки;
        - строки: qty берём из PurchaseRequestLine (в QuoteLine qty нет);
        - НЕ создаём строки, если по item уже всё обеспечено отправленными заказами (sent+).
        - пропускаем QuoteLine.is_blocked.
        """
        quote = self.get_object()

        pr = quote.purchase_request if quote.purchase_request_id else None

        # Уникальный номер заказа: PO-<quote_id>-<YYYYMMDDHHMMSS>
        po_number = f"PO-{quote.id}"

        po = PurchaseOrder.objects.create(
            supplier=quote.supplier,
            number=po_number,
            status="draft",
            quote=quote,
            purchase_request=pr,
        )

        # Считаем требуемое количество по заявке (агрегировано по item)
        required_qty = {}
        if pr is not None:
            for pr_ln in pr.lines.all():
                if pr_ln.item_id is None:
                    continue
                required_qty[pr_ln.item_id] = required_qty.get(pr_ln.item_id, Decimal("0")) + (pr_ln.qty or Decimal("0"))

        # Уже обеспечено отправленными/подтверждёнными заказами
        already_ordered = {}
        if pr is not None:
            already_ordered = _pr_ordered_qty_by_item(pr, PO_ORDERED_STATUSES)

        for ln in quote.lines.all():
            if getattr(ln, "is_blocked", False):
                continue
            item_id = ln.item_id
            if item_id is None:
                continue

            need = required_qty.get(item_id, Decimal("0"))
            have = already_ordered.get(item_id, Decimal("0"))
            remaining = need - have

            if remaining <= 0:
                continue

            PurchaseOrderLine.objects.create(
                order=po,
                item=ln.item,
                qty=remaining,
                price=ln.price,
                status="pending",
            )

        out = PurchaseOrderSerializer(po, context={"request": request})
        return Response(out.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=["post"], url_path="generate-from-request")
    @transaction.atomic
    def generate_from_request(self, request):
        """Создать КП (Quote) от выбранных поставщиков на основе заявки.

        Это контрактный эндпоинт под фронт: он НЕ требует миграций.
        На dev создаёт Quote+QuoteLine, стараясь подтянуть цену/sku из ItemSupplierMapping.

        POST /api/procurement/quotes/generate-from-request/
        {
          "purchase_request_id": 123,
          "supplier_ids": [1,2,3]
        }
        """

        serializer = GenerateQuotesFromRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        pr_id = serializer.validated_data["purchase_request_id"]
        supplier_ids = serializer.validated_data["supplier_ids"]
        line_ids = serializer.validated_data.get("purchase_request_line_ids") or []

        pr = PurchaseRequest.objects.prefetch_related("lines__item", "lines__unit").get(id=pr_id)
        pr_lines = list(pr.lines.all())
        if line_ids:
            pr_lines = [ln for ln in pr_lines if ln.id in set(line_ids)]
        if not pr_lines:
            return Response({"detail": "Заявка не содержит позиций"}, status=status.HTTP_400_BAD_REQUEST)

        suppliers = list(Supplier.objects.filter(id__in=supplier_ids))
        warnings = []
        created = []

        from procurement.serializers import _quote_meta_set

        for s in suppliers:
            q = Quote.objects.create(
                supplier=s,
                purchase_request=pr,
                source=f"RFQ from PR#{pr.id}",
            )

            # дефолтный статус для UI
            _quote_meta_set(q.id, status="received")

            for pr_ln in pr_lines:
                # Попробуем найти сопоставление item↔supplier_sku для автозаполнения
                mapping = ItemSupplierMapping.objects.filter(
                    item_id=pr_ln.item_id,
                    is_active=True,
                    price_list_line__price_list__supplier_id=s.id,
                    price_list_line__price_list__is_active=True,
                ).select_related("price_list_line", "price_list_line__unit", "price_list_line__price_list").order_by(
                    "-is_preferred", "price_list_line__price"
                ).first()

                if mapping:
                    pll = mapping.price_list_line
                    QuoteLine.objects.create(
                        quote=q,
                        item=pr_ln.item,
                        vendor_sku=pll.supplier_sku,
                        name=pll.description or pr_ln.item.name,
                        unit=pll.unit or pr_ln.unit,
                        price=pll.price,
                        currency=(getattr(pll.price_list, 'currency', None) or 'RUB'),
                        lead_days=getattr(pll, 'lead_time_days', None),
                        moq_qty=getattr(pll, 'min_quantity', None),
                        pack_qty=getattr(pll, 'package_quantity', None),
                        lot_step=getattr(pll, 'quantity_step', None),
                    )
                else:
                    warnings.append(
                        f'Нет сопоставления для "{pr_ln.item.name}" у поставщика "{s.name}" — цена = 0'
                    )
                    QuoteLine.objects.create(
                        quote=q,
                        item=pr_ln.item,
                        vendor_sku="",
                        name=pr_ln.item.name,
                        unit=pr_ln.unit,
                        price=Decimal("0"),
                        currency="RUB",
                    )

            created.append({
                "id": q.id,
                "supplier_id": s.id,
                "supplier_name": s.name,
            })

        # Для удобства фронта вернём полный список созданных КП в том же формате,
        # что и QuoteSerializer (но без тяжёлых строк — фронт при необходимости доберёт detail).
        return Response({
            "ok": True,
            "quotes_created": len(created),
            "quotes": created,
            "warnings": warnings,
        }, status=status.HTTP_201_CREATED)

class ShipmentViewSet(viewsets.ReadOnlyModelViewSet):
    """
    ViewSet для доставок, построенный на базе PurchaseOrder.

    Маршруты:
    - GET /api/procurement/shipments/
    - GET /api/procurement/shipments/{id}/
    """

    permission_classes = [permissions.IsAuthenticated]
    serializer_class = ShipmentSerializer

    def get_queryset(self):
        """
        Возвращает QuerySet PurchaseOrder, на основе которого строятся Shipment.
        """
        return PurchaseOrder.objects.select_related("supplier").order_by("-id")

    def list(self, request, *args, **kwargs):
        qs = self.get_queryset()
        data = [self._po_to_shipment(po) for po in qs]
        serializer = self.get_serializer(data, many=True)
        return Response(serializer.data)

    def retrieve(self, request, *args, **kwargs):
        po = self.get_queryset().get(pk=kwargs["pk"])
        shipment = self._po_to_shipment(po)
        serializer = self.get_serializer(shipment)
        return Response(serializer.data)

    def destroy(self, request, *args, **kwargs):
        """Удалить доставку (compat): откатывает статус заказа до 'sent'.

        В текущем MVP доставка — view-модель поверх PurchaseOrder, поэтому физически мы ничего не удаляем.
        DELETE /api/procurement/shipments/{id}/:
        - если заказ был в путях/доставлен — возвращаем в 'sent'
        - пересчитываем статус заявки (PR), т.к. статусы PO влияют на обеспечение
        """
        po = self.get_queryset().get(pk=kwargs["pk"])
        old = (po.status or "").lower()
        # откат только если уже была стадия доставки
        if old in {"in_transit", "delivered"}:
            po.status = "sent"
            po.save(update_fields=["status", "updated_at"])
            try:
                if po.purchase_request_id:
                    _recalc_purchase_request_status(po.purchase_request)
            except Exception:
                pass
        return Response(status=status.HTTP_204_NO_CONTENT)

    def _po_to_shipment(self, po: PurchaseOrder):
        """
        Маппит PurchaseOrder в dict под ShipmentSerializer.
        Поля ETA/фактической даты/трек‑номера/notes можно донастроить при появлении
        соответствующих полей в модели PurchaseOrder.
        """
        return {
            "id": po.id,
            "po_id": po.id,
            "po_number": po.number,
            "supplier_id": po.supplier_id,
            "supplier_name": po.supplier.name if po.supplier_id else "",
            "status": po.status,
            "tracking_number": "",          # пока заглушка
            "estimated_delivery": getattr(po, "deadline", None),     # дедлайн/ETA
            "actual_delivery": None,        # TODO: отдельное поле, если появится
            "notes": "",                    # TODO: поле комментария, если появится
            "created_at": po.created_at,
            "updated_at": po.updated_at,
        }


# ============================================================================
# HELPER FUNCTIONS (вспомогательные функции)
# ============================================================================

def _to_decimal(val, default=None):
    """Конвертировать значение в Decimal"""
    if val in (None, ''):
        return default
    try:
        return Decimal(str(val))
    except:
        return default


def _to_int(val, default=None):
    """Конвертировать значение в int"""
    if val in (None, ''):
        return default
    try:
        return int(val)
    except:
        return default


def _to_bool(val, default=None):
    """Конвертировать значение в bool"""
    if val is None:
        return default
    if isinstance(val, bool):
        return val
    s = str(val).strip().lower()
    if s in ('1', 'true', 'yes', 'да', 'y', 'истина', 'вкл'):
        return True
    if s in ('0', 'false', 'no', 'нет', 'n', 'ложь', 'откл'):
        return False
    return default


def resolve_item_id_by_supplier_context(supplier_name, sku):
    """
    Разрешить Item ID по контексту поставщика и SKU.
    
    Логика:
    1. Ищем ItemSupplierMapping где supplier.name совпадает и supplier_sku совпадает
    2. Если нашли — возвращаем item_id
    3. Если нет — ищем Item по SKU (по item.sku)
    4. Если ничего не нашли — возвращаем None
    
    Args:
        supplier_name: название поставщика
        sku: артикул поставщика
    
    Returns:
        int или None
    """
    if not supplier_name or not sku:
        return None
    
    try:
        supplier = Supplier.objects.get(name=supplier_name)
    except Supplier.DoesNotExist:
        supplier = None
    
    # Попытка 1: ItemSupplierMapping (если найден поставщик)
    if supplier:
        mapping = ItemSupplierMapping.objects.filter(
            price_list_line__price_list__supplier=supplier,
            price_list_line__supplier_sku=sku,
            is_active=True
        ).first()
        
        if mapping:
            return mapping.item_id
    
    # Попытка 2: Item по SKU
    item = Item.objects.filter(sku=sku).first()
    if item:
        return item.id
    
    return None


# ============================================================================
# 1. SupplierPriceListViewSet
# ============================================================================

class SupplierPriceListViewSet(viewsets.ModelViewSet):
    """
    ViewSet для управления прайс-листами поставщиков.
    
    Endpoints:
    - GET /api/procurement/supplier-price-lists/ — список
    - GET /api/procurement/supplier-price-lists/{id}/ — детали
    - POST /api/procurement/supplier-price-lists/ — создание
    - PUT /api/procurement/supplier-price-lists/{id}/ — редактирование
    - DELETE /api/procurement/supplier-price-lists/{id}/ — удаление
    - POST /api/procurement/supplier-price-lists/{id}/upload/ — загрузка из файла
    
    Query parameters:
    - supplier_id: фильтр по поставщику
    - is_active: фильтр по статусу
    - currency: фильтр по валюте
    """
    
    queryset = SupplierPriceList.objects.select_related('supplier').prefetch_related('lines')
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = (MultiPartParser, FormParser)
    
    def get_serializer_class(self):
        """Выбрать сериализатор в зависимости от действия"""
        if self.action == 'retrieve':
            return SupplierPriceListDetailSerializer
        return SupplierPriceListSerializer
    
    def get_queryset(self):
        """Фильтрация по параметрам запроса"""
        queryset = super().get_queryset()
        
        # Фильтр по поставщику
        supplier_id = self.request.query_params.get('supplier_id')
        if supplier_id:
            queryset = queryset.filter(supplier_id=supplier_id)
        
        # Фильтр по статусу
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=_to_bool(is_active))
        
        # Фильтр по валюте
        currency = self.request.query_params.get('currency')
        if currency:
            queryset = queryset.filter(currency=currency)
        
        return queryset
    
    @action(detail=True, methods=['post'], parser_classes=(MultiPartParser, FormParser))
    def upload(self, request, pk=None):
        """
        Загрузить прайс-лист из Excel файла.
        
        POST /api/procurement/supplier-price-lists/{id}/upload/
        
        Параметры:
        - file: Excel файл (.xlsx)
        - preview: 1/0 — режим предпросмотра (без сохранения в БД)
        
        Ожидаемые колонки в Excel:
        - item_sku (артикул товара)
        - description (описание товара)
        - unit (единица измерения: шт, кг, м³)
        - price (цена за единицу)
        - min_quantity (МОК)
        - quantity_step (кратность)
        - lead_time_days (сроки доставки)
        - (опционально) vat_included, vat_rate, delivery_cost_fixed и т.д.
        
        Возвращает:
        {
            "ok": true,
            "preview": true/false,
            "rows": [
                {
                    "row": 2,
                    "valid": true,
                    "supplier_sku": "KR-001",
                    "description": "Кирпич красный",
                    "price": 15.00,
                    ...
                },
                ...
            ],
            "created": 10,
            "errors": 2,
            "errors_detail": [...]
        }
        """
        
        price_list = self.get_object()
        file = request.FILES.get('file')
        preview = request.query_params.get('preview') in ('1', 'true', 'True')
        
        if not file:
            return Response(
                {'detail': 'Файл не загружен'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            # Парсинг Excel
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response(
                {'detail': f'Ошибка при чтении Excel: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Определить колонки
        headers = {}
        mapping = {
            'item_sku': 'item_sku',
            'description': 'description',
            'unit': 'unit',
            'price': 'price',
            'min_quantity': 'min_quantity',
            'quantity_step': 'quantity_step',
            'lead_time_days': 'lead_time_days',
            'vat_included': 'vat_included',
            'vat_rate': 'vat_rate',
            'delivery_cost_fixed': 'delivery_cost_fixed',
            'delivery_cost_per_unit': 'delivery_cost_per_unit',
        }
        
        for col in range(1, ws.max_column + 1):
            v = ws.cell(1, col).value
            if v is None:
                continue
            k = str(v).strip().lower()
            if k in mapping:
                headers[mapping[k]] = col
        
        if 'item_sku' not in headers or 'price' not in headers:
            return Response(
                {'detail': 'Не найдены обязательные колонки: item_sku, price'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Обработка строк
        preview_rows = []
        created = 0
        errors = 0
        errors_detail = []
        
        with transaction.atomic():
            for r in range(2, ws.max_row + 1):
                sku_raw = ws.cell(r, headers.get('item_sku', 0)).value
                if sku_raw in (None, ''):
                    continue
                
                sku = str(sku_raw).strip()
                price = _to_decimal(ws.cell(r, headers.get('price', 0)).value)
                
                row_data = {
                    'row': r,
                    'supplier_sku': sku,
                    'description': str(ws.cell(r, headers.get('description', 0)).value or ''),
                    'price': price,
                    'valid': price is not None,
                    'errors': {}
                }
                
                if not row_data['valid']:
                    row_data['errors']['price'] = 'Не найдена или некорректна'
                    errors += 1
                    errors_detail.append(f'Строка {r}: {sku} — неверная цена')
                else:
                    # Получить Unit
                    unit_name = str(ws.cell(r, headers.get('unit', 0)).value or 'шт').strip()
                    try:
                        unit = Unit.objects.get(name=unit_name)
                    except Unit.DoesNotExist:
                        row_data['errors']['unit'] = f'Единица "{unit_name}" не найдена'
                        errors += 1
                        errors_detail.append(f'Строка {r}: {sku} — неверная единица {unit_name}')
                        row_data['valid'] = False
                    
                    if row_data['valid']:
                        # Создать SupplierPriceListLine
                        if not preview:
                            try:
                                line = SupplierPriceListLine.objects.create(
                                    price_list=price_list,
                                    supplier_sku=sku,
                                    description=row_data['description'],
                                    unit=unit,
                                    price=price,
                                    min_quantity=_to_decimal(
                                        ws.cell(r, headers.get('min_quantity', 0)).value, Decimal('1')
                                    ),
                                    quantity_step=_to_decimal(
                                        ws.cell(r, headers.get('quantity_step', 0)).value, Decimal('1')
                                    ),
                                    lead_time_days=_to_int(
                                        ws.cell(r, headers.get('lead_time_days', 0)).value, 14
                                    ),
                                    vat_included=_to_bool(
                                        ws.cell(r, headers.get('vat_included', 0)).value, False
                                    ),
                                    vat_rate=_to_decimal(
                                        ws.cell(r, headers.get('vat_rate', 0)).value, Decimal('20')
                                    ),
                                    delivery_cost_fixed=_to_decimal(
                                        ws.cell(r, headers.get('delivery_cost_fixed', 0)).value, Decimal('0')
                                    ),
                                    delivery_cost_per_unit=_to_decimal(
                                        ws.cell(r, headers.get('delivery_cost_per_unit', 0)).value, Decimal('0')
                                    ),
                                )
                                created += 1
                            except Exception as e:
                                row_data['errors']['create'] = str(e)
                                row_data['valid'] = False
                                errors += 1
                                errors_detail.append(f'Строка {r}: {sku} — {str(e)}')
                
                preview_rows.append(row_data)
        
        return Response({
            'ok': True,
            'preview': preview,
            'price_list_id': price_list.id,
            'rows': preview_rows,
            'created': created,
            'errors': errors,
            'errors_detail': errors_detail,
        }, status=status.HTTP_200_OK)
    
    @action(detail=False, methods=['get'])
    def active(self, request):
        """Получить список активных прайс-листов"""
        queryset = self.get_queryset().filter(is_active=True)
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data)


# ============================================================================
# 2. SupplierPriceListLineViewSet
# ============================================================================

class SupplierPriceListLineViewSet(viewsets.ModelViewSet):
    """
    ViewSet для управления позициями в прайс-листах.
    
    Endpoints:
    - GET /api/procurement/price-list-lines/
    - GET /api/procurement/price-list-lines/{id}/
    - POST /api/procurement/price-list-lines/
    - PUT /api/procurement/price-list-lines/{id}/
    - DELETE /api/procurement/price-list-lines/{id}/
    
    Query parameters:
    - price_list_id: фильтр по прайс-листу
    - supplier_id: фильтр по поставщику
    - is_available: фильтр по доступности
    """
    
    queryset = SupplierPriceListLine.objects.select_related('price_list', 'unit')
    serializer_class = SupplierPriceListLineSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return SupplierPriceListLineDetailSerializer
        return SupplierPriceListLineSerializer
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Фильтр по прайс-листу
        price_list_id = self.request.query_params.get('price_list_id')
        if price_list_id:
            queryset = queryset.filter(price_list_id=price_list_id)
        
        # Фильтр по поставщику
        supplier_id = self.request.query_params.get('supplier_id')
        if supplier_id:
            queryset = queryset.filter(price_list__supplier_id=supplier_id)
        
        # Фильтр по доступности
        is_available = self.request.query_params.get('is_available')
        if is_available is not None:
            queryset = queryset.filter(is_available=_to_bool(is_available))
        
        # Фильтр по поиску
        search = self.request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(supplier_sku__icontains=search) |
                Q(description__icontains=search)
            )
        
        return queryset


# ============================================================================
# 3. ItemSupplierMappingViewSet
# ============================================================================

class ItemSupplierMappingViewSet(viewsets.ModelViewSet):
    """
    ViewSet для управления сопоставлениями Item ↔ Поставщик.
    
    Endpoints:
    - GET /api/procurement/item-supplier-mappings/
    - GET /api/procurement/item-supplier-mappings/{id}/
    - POST /api/procurement/item-supplier-mappings/
    - PUT /api/procurement/item-supplier-mappings/{id}/
    - DELETE /api/procurement/item-supplier-mappings/{id}/
    - GET /api/procurement/item-supplier-mappings/find-alternative/
    
    Query parameters:
    - item_id: фильтр по Item
    - supplier_id: фильтр по поставщику
    - is_active: фильтр по статусу
    - is_preferred: фильтр по предпочтению
    """
    
    queryset = ItemSupplierMapping.objects.select_related(
        'item',
        'price_list_line',
        'price_list_line__price_list',
        'price_list_line__price_list__supplier'
    )
    serializer_class = ItemSupplierMappingSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # Фильтр по Item
        item_id = self.request.query_params.get('item_id')
        if item_id:
            queryset = queryset.filter(item_id=item_id)
        
        # Фильтр по поставщику
        supplier_id = self.request.query_params.get('supplier_id')
        if supplier_id:
            queryset = queryset.filter(price_list_line__price_list__supplier_id=supplier_id)
        
        # Фильтр по статусу
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=_to_bool(is_active))
        
        # Фильтр по предпочтению
        is_preferred = self.request.query_params.get('is_preferred')
        if is_preferred is not None:
            queryset = queryset.filter(is_preferred=_to_bool(is_preferred))
        
        return queryset
    
    @action(detail=False, methods=['get'])
    def find_alternative(self, request):
        """
        Найти альтернативных поставщиков для Item'а.
        
        GET /api/procurement/item-supplier-mappings/find-alternative/
        
        Query parameters:
        - item_id: ID Item'а
        - exclude_supplier_id: (опционально) ID поставщика, которого исключить
        
        Возвращает:
        {
            "item_id": 1,
            "item_name": "Кирпич красный",
            "suppliers": [
                {
                    "id": 1,
                    "supplier_name": "ООО Стройсервис",
                    "supplier_sku": "KR-001",
                    "price": 15.00,
                    "is_preferred": true,
                    ...
                },
                ...
            ]
        }
        """
        
        item_id = request.query_params.get('item_id')
        exclude_supplier_id = request.query_params.get('exclude_supplier_id')
        
        if not item_id:
            return Response(
                {'detail': 'Параметр item_id обязателен'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            item = Item.objects.get(id=item_id)
        except Item.DoesNotExist:
            return Response(
                {'detail': f'Item с ID {item_id} не найден'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Получить все активные сопоставления
        mappings = ItemSupplierMapping.objects.filter(
            item_id=item_id,
            is_active=True
        ).select_related(
            'price_list_line',
            'price_list_line__price_list',
            'price_list_line__price_list__supplier'
        ).order_by('-is_preferred', 'price_list_line__price')
        
        # Исключить поставщика если указан
        if exclude_supplier_id:
            mappings = mappings.exclude(
                price_list_line__price_list__supplier_id=exclude_supplier_id
            )
        
        data = {
            'item_id': item.id,
            'item_name': item.name,
            'item_sku': item.sku,
            'suppliers': ItemSupplierMappingSerializer(mappings, many=True).data
        }
        
        return Response(data)


# ============================================================================
# 4. GENERATE QUOTES (Автогенерация КП)
# ============================================================================

@action(detail=False, methods=['post'])
def generate_quotes_from_request(self, request):
    """
    Автоматически создать КП от поставщиков на основе заявки.
    
    POST /api/procurement/quotes/generate-from-request/
    
    Body:
    {
        "purchase_request_id": 123,
        "supplier_ids": [1, 2, 3]
    }
    
    Алгоритм:
    1. Получить заявку и её позиции
    2. Для каждого поставщика:
        a. Создать новую Quote
        b. Для каждой позиции заявки:
            - Найти ItemSupplierMapping
            - Рассчитать количество с conversion_factor
            - Создать QuoteLine
    3. Вернуть список созданных КП с warnings
    
    Возвращает:
    {
        "ok": true,
        "quotes_created": 3,
        "quotes": [
            {
                "id": 1001,
                "supplier_id": 1,
                "supplier_name": "ООО Стройсервис",
                "total_price": 150000,
                "lines_count": 2,
                "status": "new"
            },
            ...
        ],
        "warnings": [
            "Позиция Item 2 не найдена в прайс-листе поставщика 2"
        ]
    }
    """
    
    serializer = GenerateQuotesFromRequestSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    
    purchase_request_id = serializer.validated_data['purchase_request_id']
    supplier_ids = serializer.validated_data['supplier_ids']
    
    # Получить заявку
    try:
        pr = PurchaseRequest.objects.prefetch_related('lines__item').get(id=purchase_request_id)
    except PurchaseRequest.DoesNotExist:
        return Response(
            {'detail': f'PurchaseRequest {purchase_request_id} не найдена'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    if not pr.lines.exists():
        return Response(
            {'detail': 'Заявка не содержит позиций'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Получить поставщиков
    suppliers = Supplier.objects.filter(id__in=supplier_ids)
    
    quotes_created = []
    warnings = []
    
    with transaction.atomic():
        for supplier in suppliers:
            quote = Quote.objects.create(
                purchase_request=pr,
                supplier=supplier,
                status='new',
                currency='RUB'  # TODO: взять из прайс-листа
            )
            
            lines_created = 0
            
            # Для каждой позиции заявки
            for pr_line in pr.lines.all():
                # Найти сопоставление
                mapping = ItemSupplierMapping.objects.filter(
                    item_id=pr_line.item_id,
                    is_active=True
                ).select_related(
                    'price_list_line',
                    'price_list_line__price_list'
                ).filter(
                    price_list_line__price_list__supplier=supplier,
                    price_list_line__price_list__is_active=True
                ).first()
                
                if not mapping:
                    warnings.append(
                        f'Позиция "{pr_line.item.name}" не найдена в прайс-листе '
                        f'поставщика "{supplier.name}"'
                    )
                    continue
                
                # Рассчитать количество в единицах поставщика
                supplier_qty = mapping.convert_item_qty_to_supplier_qty(pr_line.quantity)
                
                # Проверить МОК
                min_qty = mapping.get_effective_min_quantity()
                if pr_line.quantity < min_qty:
                    warnings.append(
                        f'Позиция "{pr_line.item.name}" (заказано {pr_line.quantity}) '
                        f'меньше МОК {min_qty} у поставщика "{supplier.name}"'
                    )
                    # Всё равно создаём QuoteLine, но отмечаем в notes
                
                # Создать QuoteLine
                line_price = mapping.price_list_line.price
                line_total = supplier_qty * line_price
                
                quote_line = QuoteLine.objects.create(
                    quote=quote,
                    item=pr_line.item,
                    quantity=supplier_qty,
                    unit=mapping.price_list_line.unit,
                    unit_price=line_price,
                    total_price=line_total,
                    notes=f'Коэффициент преобразования: {mapping.conversion_factor}'
                )
                
                lines_created += 1
            
            if lines_created > 0:
                quotes_created.append({
                    'id': quote.id,
                    'supplier_id': supplier.id,
                    'supplier_name': supplier.name,
                    'total_price': quote.get_total_price(),
                    'lines_count': lines_created,
                    'status': quote.status
                })
            else:
                # Удалить пустую КП
                quote.delete()
    
    return Response({
        'ok': True,
        'quotes_created': len(quotes_created),
        'quotes': quotes_created,
        'warnings': warnings
    }, status=status.HTTP_201_CREATED)
