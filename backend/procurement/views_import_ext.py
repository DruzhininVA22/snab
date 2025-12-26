"""
Расширенный импорт/предпросмотр документов закупки.

Содержит ручки, которые помогают загрузить и предварительно разобрать документы (счета/накладные),
чтобы затем сопоставить их с заявками, заказами и номенклатурой.
"""

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from django.db import transaction
from catalog.models import Item
from suppliers.models import Supplier
from procurement.models import PriceRecord
from datetime import date
import openpyxl, mimetypes


from .importers.pdf_price import parse_pdf as parse_price_pdf


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def import_invoice_preview(request):
    """
    Предпросмотр импорта счёта/накладной: разбирает файл и возвращает структуру строк.
    Используется для сверки позиций счёта с расчётами/заявками до фактической загрузки.
    """

    file = request.FILES.get("file")
    supplier = request.data.get("supplier") or request.query_params.get("supplier") or ""
    if not file:
        return Response({"detail": "Нет файла"}, status=400)
    ctype = (file.content_type or mimetypes.guess_type(file.name)[0] or "").lower()
    if "pdf" in ctype or file.name.lower().endswith(".pdf"):
        # Для счётов используем тот же базовый PDF-парсер как превью (SKU / name / price)
        parsed = parse_price_pdf(file, supplier_name=supplier)
        parsed["document_type"] = "invoice"
        return Response(parsed, status=200 if parsed.get("ok") else 400)
    # Excel-счёт: используем простую схему — те же заголовки, что и прайс
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception as e:
        return Response({"detail": f"Excel error: {e}"}, status=400)
    headers = {}
    mapping = {
        "item_sku":"item_sku", "supplier":"supplier", "price":"price", "currency":"currency",
        "lead_days":"lead_days", "pack_qty":"pack_qty", "moq_qty":"moq_qty", "mo_amount":"mo_amount",
        "lot_step":"lot_step", "dt":"dt",
    }
    for col in range(1, ws.max_column+1):
        v = ws.cell(1, col).value
        if v is None: continue
        k = str(v).strip().lower()
        if k in mapping:
            headers[mapping[k]] = col
    rows = []
    for r in range(2, ws.max_row+1):
        if "item_sku" not in headers: break
        sku = ws.cell(r, headers["item_sku"]).value
        if sku in (None,""): continue
        price = ws.cell(r, headers.get("price",0)).value if "price" in headers else None
        rows.append({
            "row": r, "valid": price is not None,
            "item_sku": str(sku), "supplier": supplier or (ws.cell(r, headers.get("supplier",0)).value if "supplier" in headers else ""),
            "price": price, "currency": (ws.cell(r, headers.get("currency",0)).value if "currency" in headers else "RUB"),
        })
    return Response({"ok": True, "preview": True, "document_type":"invoice", "rows": rows}, status=200)